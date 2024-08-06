import os
import shutil

import openpyxl
from src.context import Context

def generate_diagram():
    network_design_fname = Context.active_unit + '/network_design.xlsx'
    network_diagram_fname = Context.active_unit + '/network_diagram'
    gen_widgets_db([], network_design_fname, network_diagram_fname)
    os.system('start excel "' + os.getcwd() + '/' + network_diagram_fname + '.xlsm"')
    #os.system('start excel "' + os.getcwd() + '/' + network_diagram_fname + '.xlsx"')
    pass

def gen_widgets_db(pidrozdily, network_design_fname, network_diagram_fname):
    #build relations
    wb = openpyxl.load_workbook(filename=network_design_fname)
    ws = wb.active
    ws_rows_count = len(list(ws.rows))
    ws_cols_count = len(list(ws.columns))
    devices = []
    nets = {}
    p_used = set()
    if not pidrozdily:
        pidrozdily = [
            'бригада1',
            'бат1', 
            'бат2',
            'бат3',
            'рота1', 
            'рота2', 
            'рота3', 
            'взвод1',
            'взвод2', 
            'взвод3', 
            'відділення1',
            'відділення2', 
            'відділення3', 
            'взвод КВ',
            'взвод ГВ',
            'взвод ЗРВ',
            'взвод РВ',
            'взвод ПТВ',
            'взвод ІСВ',
            'взвод ВМЗ',
            'взвод ВТЗ',
            'взвод ВЗ',
            'взвод МП'
            ]
        #for p in range(1,100):
        #    pidrozdily = pidrozdily + [True]
    for pidrozdil in pidrozdily:
        for r in range(2,ws_rows_count): #read from 3 (start) line 
            #print('row ' + str(r))
            in_pidrozdil = False
            c = 1
            # to Function column
            #if pidrozdil == True:
            #    pidrozdil = ws.cell(row=r, column=c).value
            while c <= ws_cols_count: 
                if ws.cell(row=r, column=c).value == pidrozdil and (not ws.cell(row=r, column=c+1).value or ws.cell(row=1, column=c+1).value == 'Функція'):
                    in_pidrozdil = True
                c = c + 1
                if ws.cell(row=1, column=c).value == 'Функція':
                    break
            if not in_pidrozdil:
                continue
            c = c + 1
            user_name = ''
            while c <= ws_cols_count:
                #print('col ' + str(c))
                if ws.cell(row=1, column=c).value == 'Назва':
                    user_name = ws.cell(row=r, column=c).value
                    # skip Радіостанція column
                    c = c + 2
                    #print(user_name + ' in pidrozdil ' + pidrozdil)
                    #print(user_name + '...')
                    continue
                if ws.cell(row=r, column=c).value:
                    net = ws.cell(row=1, column=c).value
                    dev = next(filter(lambda d: d['pidrozdil'] == pidrozdil and d['net'] == net, devices), {})
                    if dev:
                        devices.remove(dev)
                        if not user_name in dev['user_name']:
                            dev['user_name'] = dev['user_name'] + ', ' + user_name
                    else:
                        dev['user_name'] = user_name
                    dev['net'] = net.replace('RadioNet','')
                    nets[dev['net']] = ''
                    dev['model_name'] = ws.cell(row=r, column=8).value
                    dev['pidrozdil'] = pidrozdil
                    p_used.add(pidrozdil)
                    devices = devices + [dev]
                    #print(len(devices))
                c = c + 1
    #print(devices)
    #print(nets)
    #exit()
    wb.close()
    #exit()
    # end build 
    dev_w = 118
    col_span = 20
    RIGHT = len(p_used)*(col_span + dev_w)
    TOP = 10
    flag_height = 80
    flag_width = 40
    flag_span = 80
    conn_span = 10
    dev_height = 45
    dev_width = 118
    net_to_dev_offset = -5
    network_diagram_template_fname = 'src/network_diagram.xlsm'
    shutil.copyfile(network_diagram_template_fname, network_diagram_fname + '.xlsm')
    #wb = openpyxl.load_workbook(filename=network_diagram_fname, read_only=False, keep_vba=True)
    #ws = wb['widgets_db']
    wb = openpyxl.Workbook()
    ws = wb.active
    ws_rows_count = len(list(ws.rows))
    ws_cols_count = len(list(ws.columns))
    r = 2
    p_i = 0
    p_prev = ''
    net_i = 0
    net_prev = ''
    nets_widgets = {}
    flag_widgets = {}
    for device in devices:
        if not p_prev == device['pidrozdil']:
            p_prev = device['pidrozdil']
            p_i = p_i + 1
        net_i = list(nets.keys()).index(device['net'])
        #if not net_prev == device['net']:
        #    net_prev = device['net']
        #    net_i = net_i + 1
        #ws.cell(row=r, column=1, value=device['user_name'].split('_')[0][:-1])
        ws.cell(row=r, column=1, value=device['user_name'])
        #ws.cell(row=r, column=1, value=device['net'])
        ws.cell(row=r, column=2, value=device['model_name'])
        #ws.cell(row=r, column=3, value=device['pidrozdil'])
        ws.cell(row=r, column=3, value=get_device_trait(device['model_name'], 'connection_type'))
        left = RIGHT - dev_w - (p_i-1)*(col_span + dev_w)
        flag_widgets.setdefault(device['pidrozdil'], {'left':left + dev_w/2 - flag_width/2, 'right':0, 'top':TOP})
        ws.cell(row=r, column=4, value=left)
        dev_h = TOP + flag_height + flag_span + (net_i - 1)*(conn_span + dev_height)
        net_h = TOP + flag_height + flag_span + (net_i)*(conn_span + dev_height) - conn_span + net_to_dev_offset
        net_name = device['net']
        nets_widgets.setdefault(net_name, {'left':10000, 'right':0})
        nets_widgets[net_name]['top'] = net_h
        #nets_widgets[net_name]['left'] = min(nets_widgets[net_name]['left'], left + dev_width/2 + net_to_dev_offset)
        #nets_widgets[net_name]['right'] = max(nets_widgets[net_name]['right'], left + dev_width/2 + net_to_dev_offset)
        nets_widgets[net_name]['left'] = min(nets_widgets[net_name]['left'], left)
        nets_widgets[net_name]['right'] = max(nets_widgets[net_name]['right'], left + dev_width)
        ws.cell(row=r, column=5, value=dev_h)
        r = r + 1
    r_offset = r+2
    for r,net_name in enumerate(nets_widgets.keys()):
        #ws.cell(row=r, column=1, value=net_name)
        ws.cell(row=r+r_offset, column=1, value='solid_line')
        ws.cell(row=r+r_offset, column=2, value=nets_widgets[net_name]['left'])
        ws.cell(row=r+r_offset, column=3, value=nets_widgets[net_name]['right'])
        ws.cell(row=r+r_offset, column=4, value=nets_widgets[net_name]['top'])
    r_offset = r_offset + r + 1
    for r,p_name in enumerate(flag_widgets.keys()):
        ws.cell(row=r+r_offset, column=2, value=flag_widgets[p_name]['left'])
        ws.cell(row=r+r_offset, column=3, value=flag_widgets[p_name]['right'])
        ws.cell(row=r+r_offset, column=4, value=flag_widgets[p_name]['top'])
        if p_name.startswith('бат'):
            ws.cell(row=r+r_offset, column=1, value='bat_flag')
        elif p_name.startswith('бригада'):
            ws.cell(row=r+r_offset, column=1, value='birg_flag')
        elif p_name.startswith('рота'):
            ws.cell(row=r+r_offset, column=1, value='rota_flag')
        elif p_name.startswith('взвод'):
            ws.cell(row=r+r_offset, column=1, value='vzvod_flag')
            ws.cell(row=r+r_offset, column=4, value=flag_widgets[p_name]['top']+5)
        elif p_name.startswith('відділ'):
            ws.cell(row=r+r_offset, column=1, value='viddil_flag')
        # unit frame
        r_offset = r_offset + 1
        ws.cell(row=r+r_offset, column=1, value='unit_frame')
        ws.cell(row=r+r_offset, column=2, value=flag_widgets[p_name]['left']  + flag_width/2 - dev_w/2)
        ws.cell(row=r+r_offset, column=3, value=flag_widgets[p_name]['right'])
        ws.cell(row=r+r_offset, column=4, value=flag_widgets[p_name]['top'] + flag_height)
        # unit header
        r_offset = r_offset + 1
        ws.cell(row=r+r_offset, column=1, value='unit_header')
        ws.cell(row=r+r_offset, column=2, value=flag_widgets[p_name]['left']  + flag_width/2 - dev_w/2)
        ws.cell(row=r+r_offset, column=3, value=flag_widgets[p_name]['right'])
        ws.cell(row=r+r_offset, column=4, value=flag_widgets[p_name]['top'] + flag_height)
        ws.cell(row=r+r_offset, column=5, value=p_name)
    #print('Saving ' + network_diagram_fname + '.xlsx')
    wb.save(filename=network_diagram_fname + '.xlsx')
    wb.close()

def get_device_trait(dev_model, trait_name):
    if trait_name == 'connection_type':
        if dev_model.startswith('DP ') or dev_model.startswith('DM '):
            return 'ТрЗ'
    pass