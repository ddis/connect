import os

import openpyxl
from src.helper import collect_tree_keys
from src.context import Context, Globals
from src.ui import update_view

# Generate file network_design.xlsx
def create_network_design(fname):
    all_keys = {}
    collect_tree_keys(Globals.organizational_design, all_keys)
    # all_keys.sort(key=lambda x: len(x['units']))
    all_keys = dict(sorted(all_keys.items(), key=lambda x: len(x[1]["units"])))
    # print(all_keys.keys())
    # exit()
    wb = openpyxl.Workbook()
    ws = wb.active

    headers = [
        "Бригада",
        "Батальон",
        "Рота",
        "Взвод",
        "Відділення",
        "Функція",
        "Назва",
        "Радіостанція",
    ]

    for column, value in enumerate(headers):
        ws.cell(row=1, column=column + 1, value=value)

    r = 2
    for unique_user_name in all_keys:
        user_name = all_keys[unique_user_name]["user_name"]
        org = all_keys[unique_user_name]["units"]
        com_devices_def_count = 3
        for dev_i in range(1, com_devices_def_count + 1):
            for p in org:
                if "бригада" in p:
                    ws.cell(row=r, column=1, value=p)
                if "бат" in p:
                    ws.cell(row=r, column=2, value=p)
                if "рота" in p:
                    ws.cell(row=r, column=3, value=p)
                if "взвод" in p:
                    ws.cell(row=r, column=4, value=p)
                if "відділ" in p:
                    ws.cell(row=r, column=5, value=p)
                func = ""
                if user_name.startswith("ком"):
                    # print(user_name)
                    func = "commander"
                if user_name.startswith("GolSer"):
                    func = "serzhant"
                if user_name.startswith("Sniper"):
                    func = "sniper"
            # ws.cell(row=r, column=6, value=func)
            ws.cell(row=r, column=7, value=user_name)
            if dev_i == 1:
                ws.cell(row=r, column=8, value="DP 4400")

            if checkUnitRole(user_name):
                if dev_i == 1:
                    ws.cell(row=r, column=12, value="Starlink")
                    pass
                if dev_i == 2:
                    ws.cell(row=r, column=12, value="TooWay")
                    pass
            r += 1
            if not func == "commander":
                break

    std_networks = [
        #'ComBrigOpenPhoneNet(Farlep)?',
        #'ComBrigSecurePhoneNet(MOSI)?',
        #'ComBrigRadioNet',
        #'ComBrigDataNet?OrZSUDataNets?',
        #'ComBatOpenPhoneNet(Farlep)?',
        #'ComBatSecurePhoneNet(MOSI)?',
        "комбат мр",
        #'ComBatDataNet?OrZSUDataNets?',
        "комроти мр",
        "комвзв мр",
        #'MOSIPhoneNet?',
        #'ZSU002DataNet',
        #'ZSU001DataNet',
        #'WorldDataNet(Internet)',
        #'WorldPhoneNet'
    ]
    for i, std_net in enumerate(std_networks):
        ws.cell(row=1, column=i + (len(headers) + 1), value=std_net)

    column_position = len(headers) + len(std_networks) + 1

    ws.cell(row=1, column=column_position, value="Супутниковий зв’язок")

    column_position += 1 

    satellite_devices = [
        "Телефон",
        "Internet",
        "Передача данних",
        "bla-bla"
    ]

    for i, satellite_device in enumerate(satellite_devices):
        ws.cell(row=1, column=column_position + i, value=satellite_device)

    add_satellite_devices(ws)

    wb.save(fname)
    print("Saved " + fname)

# Add satellites into units
def add_satellite_devices(ws):
    
    pass

# Set-up config and relations for devises in unit
def setup_nets():
    network_design_fname = Context.active_unit + "/network_design.xlsx"
    if not os.path.isfile(network_design_fname):
        print(network_design_fname + "is not detected, creating template")
        create_network_design(network_design_fname)
    # QMessageBox.warning(None, '', 'Remove filter view before exit')
    print("Opening " + network_design_fname)
    # subprocess.call('EXCEL.EXE "' + os.getcwd() + '/' + network_design_fname + '"', shell=True)
    os.system(
        'start /WAIT EXCEL.EXE "' + os.getcwd() + "/" + network_design_fname + '"'
    )

    print("Updating in-memory data")
    org_db_new = {}
    nets_design_db_new = {}
    nets_config_db_new = {}
    devs_config_db_new = {}
    wb = openpyxl.load_workbook(network_design_fname)
    ws = wb.active
    dev_i = 1
    prev_user_name = ""
    for r in range(2, len(list(ws.rows)) + 1):
        user_name = ws.cell(row=r, column=7).value
        dev_model = ws.cell(row=r, column=8).value
        if not dev_model:
            continue
        if not user_name == prev_user_name:
            dev_i = 1
            prev_user_name = user_name
            org_db_new[len(org_db_new)] = user_name
        dev_name = user_name + " рст" + str(dev_i)
        # update org_db with unit path
        is_in_some_net = False
        for c in range(9, len(list(ws.columns)) + 1):
            is_in_net = ws.cell(row=r, column=c).value
            is_in_some_net = is_in_some_net or is_in_net
            if is_in_net:
                nets_design_db_new.setdefault(user_name, {})
                net_name = ws.cell(row=1, column=c).value
                print(dev_name)
                nets_design_db_new[user_name][net_name] = dev_name
                nets_config_db_new.setdefault(
                    net_name,
                    {
                        "net_name": net_name,
                        "channel": len(nets_config_db_new) + 1,
                        "caller_id": int.from_bytes(os.urandom(2), "big"),
                        "Rx": 0,
                        "Tx": 0,
                        "Timeslot": "1",
                        "ColorCode": "7",
                        "key_id": int.from_bytes(os.urandom(1), "big"),
                        "key_value": "",
                    },
                )
                devs_config_db_new.setdefault(
                    dev_name,
                    {
                        "dev_name": dev_name,
                        "model_name": dev_model,
                        "nets_names": [],
                        "firmware_fname": "",
                    },
                )
                devs_config_db_new[dev_name]["nets_names"] = devs_config_db_new[
                    dev_name
                ]["nets_names"] + [net_name]
                print(user_name + "-" + dev_name)
        if is_in_some_net:
            dev_i = dev_i + 1
    if Context.nets_design_db == nets_design_db_new:
        print("nets_design_db is the same, update is not needed")
        return
    Context.nets_design_db = nets_design_db_new
    Context.org_db = org_db_new
    print()
    print("nets_design_db updated")
    # reset firmware always when network design is changed
    Context.devs_config_db = devs_config_db_new
    print("devs_config_db updated")
    # reset nets_config_db if nets list was changed
    if not Context.nets_config_db.keys() == nets_config_db_new.keys():
        Context.nets_config_db = nets_config_db_new
        print("nets_config_db updated")
    update_view()


def configure_nets():
    nets_config_fname = Context.active_unit + "/network_config.xlsx"

    # autofill empty parameters for channels
    def generate_net_param(param_name, v):

        if param_name == "key_value":
            return generate_key()
        elif param_name == "Rx":
            already_used = [float(v["Rx"]) for v in Context.nets_config_db.values()]
            Rx = get_next_available_Freq(0)
            while Rx in list(already_used):
                Rx = get_next_available_Freq(Rx)
            return Rx
        elif param_name == "Tx":
            return v["Rx"]
        pass

    # create excel file from nets_config_db, open it, wait when editing is finished
    wb = openpyxl.Workbook()
    ws = wb.active
    for r, net in enumerate(Context.nets_config_db.values()):
        for c, net_param_name in enumerate(net.keys()):
            ws.cell(row=1, column=c + 1, value=net_param_name)
            if not net[net_param_name]:
                net[net_param_name] = generate_net_param(net_param_name, net)
                print("generated: " + str(net[net_param_name]))
            ws.cell(row=r + 2, column=c + 1, value=net[net_param_name])
    wb.save(nets_config_fname)
    os.system('start /WAIT EXCEL.EXE "' + os.getcwd() + "/" + nets_config_fname + '"')
    # load nets_config_db from excel file (save to json file in update_view)
    wb = openpyxl.load_workbook(nets_config_fname)
    ws = wb.active
    nets_config_db_new = {}
    for r, row in enumerate(ws.rows):
        if r == 0:
            continue
        net_name = ws.cell(row=r + 1, column=1).value
        nets_config_db_new[net_name] = {}
        for c, col in enumerate(ws.columns):
            nets_config_db_new[net_name][ws.cell(row=1, column=c + 1).value] = ws.cell(
                row=r + 1, column=c + 1
            ).value
    print(nets_config_db_new)
    # reset devs config
    if not nets_config_db_new == Context.nets_config_db:
        Context.nets_config_db = nets_config_db_new
        for dev in Context.devs_config_db.values():
            dev["firmware_fname"] = ""
    update_view()


def generate_key():
    # 128 bit
    return os.urandom(16).hex()


def get_next_available_Freq(Rx):
    step = 0.5
    if not Rx:
        Rx = 140.0 - step
    return Rx + step

def checkUnitRole(user_name: str) -> bool:
    return user_name.startswith("комбриг") or user_name.startswith("комбат") or user_name.startswith("комроти")
